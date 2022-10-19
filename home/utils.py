import site
import openpyxl
from pathlib import Path
from django.contrib.gis.geos import Point

from district.models import District, Village
from home.models import Chill, ChillType, Media, TouristSites


def read_data(file):
    wb_obj = openpyxl.load_workbook(file)
    sheet = wb_obj.active

    sites = []

    last_column = len(list(sheet.columns))
    last_row = len(list(sheet.rows))

    for row in range(1, last_row + 1):
        building = {}
        for column in range(1, last_column + 1):
            column_letter = openpyxl.utils.get_column_letter(column)
            if row > 1:
                building[sheet[column_letter + str(1)].value] = sheet[
                    column_letter + str(row)
                ].value
                # print(building)
        sites.append(building)

    sites.pop(0)

    for rw in sites:
        village, _ = Village.objects.get_or_create(name=rw.get("name"))
        loc = Point(float(rw.get("longitude")), float(rw.get("latitude")))
        district = District.objects.get(geom__intersects=loc)
        print(rw)
        tour_site, _ = TouristSites.objects.get_or_create(
            name=rw.get("name"),
            year_of_inception=rw.get("inception"),
            fee=rw.get("fee"),
            phone=rw.get("phone"),
            popular_stop_overs=rw.get("Popular_Stop_Overs"),
            nearest_main_hospital=rw.get("main_hospital"),
            location=loc,
            district=district,
            village=village
        )
        links = []
        try:
            link = str(rw.get("image"))
            links.append(link)
        except:
            pass
        try:
            link2 = rw.get("image_2")
            links.append(link2)
        except:
            pass
        try:
            link3 = rw.get("image_3")
            links.append(link3)
        except:
            pass
        for link in links:
            try:
                Media.objects.get_or_create(link=link, chill=tour_site)
            except:
                pass
        print("Added")


def chill_data(file):
    wb_obj = openpyxl.load_workbook(file)
    sheet = wb_obj.active

    sites = []

    last_column = len(list(sheet.columns))
    last_row = len(list(sheet.rows))

    for row in range(1, last_row + 1):
        building = {}
        for column in range(1, last_column + 1):
            column_letter = openpyxl.utils.get_column_letter(column)
            if row > 1:
                building[sheet[column_letter + str(1)].value] = sheet[
                    column_letter + str(row)
                ].value
                # print(building)
        sites.append(building)

    sites.pop(0)

    for rw in sites:
        type, _ = ChillType.objects.get_or_create(name=rw.get("type"))
        loc = Point(float(rw.get("longitude")), float(rw.get("latitude")))
        district = District.objects.get(geom__intersects=loc)
        print(rw)
        print(district)
        chill_site, _ = Chill.objects.get_or_create(
            name=rw.get("name"),
            contact=rw.get("inception"),
            fee=rw.get("fee"),
            checkin=rw.get("checkin"),
            checkout=rw.get("checkout"),
            dishes=rw.get("dishes"),
            foods=rw.get("foods"),
            sauces=rw.get("sauces"),
            district=district,
            website=rw.get("website"),
            email=rw.get("email"),
            type=type,
            number_of_beds=rw.get("number_of_beds"),
            number_of_rooms=rw.get("number_of_bedrooms"),
            location=loc

        )
        links = []
        try:
            link = str(rw.get("image"))
            links.append(link)
        except:
            pass
        try:
            link2 = rw.get("image_2")
            links.append(link2)
        except:
            pass
        try:
            link3 = rw.get("image_3")
            links.append(link3)
        except:
            pass
        for link in links:
            try:
                Media.objects.get_or_create(link=link, chill=chill_site)
            except:
                pass
        print("Added")
